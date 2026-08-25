"""Extract the Positions/Orders/Deals sections of ReportHistory-901018.xlsx
into CSVs consumed by the rest of the parity toolkit.

Usage:
    python extract_report.py [--xlsx ReportHistory-901018.xlsx] [--out out/]

Produces: out/positions.csv, out/orders.csv, out/deals.csv (raw report rows).
"""
from __future__ import annotations

import argparse
import csv
import os
import warnings

SECTIONS = {"Positions", "Orders", "Deals", "Open Positions", "Working Orders", "Results"}


def extract(xlsx_path: str, out_dir: str) -> dict[str, int]:
    import openpyxl  # local import: only needed for extraction

    warnings.filterwarnings("ignore")
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    section = None
    expect_header = False
    writers: dict[str, csv.writer] = {}
    files = {}
    counts: dict[str, int] = {}

    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if isinstance(first, str) and first.strip() in SECTIONS and all(v is None for v in row[1:]):
            section = first.strip()
            expect_header = True
            continue
        if expect_header and section is not None:
            fn = os.path.join(out_dir, section.lower().replace(" ", "_") + ".csv")
            f = open(fn, "w", newline="")
            w = csv.writer(f)
            w.writerow([str(c) for c in row])
            writers[section] = w
            files[section] = f
            counts[section] = 0
            expect_header = False
            continue
        if section in writers and row[0] is not None:
            writers[section].writerow(row)
            counts[section] += 1

    for f in files.values():
        f.close()
    return counts


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="ReportHistory-901018.xlsx")
    ap.add_argument("--out", default="tools/parity_sim/out")
    args = ap.parse_args()
    counts = extract(args.xlsx, args.out)
    print("extracted:", counts)
